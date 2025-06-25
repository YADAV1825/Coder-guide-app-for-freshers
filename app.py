#*
# MADE BY: ROHIT YADAV
#          NIT JALANDHAR
#          2023-2027*#


import tkinter as tk
from tkinter import messagebox
import json
import os
import webbrowser
import pandas as pd
import platform
from openpyxl import load_workbook 
import os
import sys
import json

def get_data_file_path(filename):
    """Get the correct path for bundled JSON files."""
    if getattr(sys, 'frozen', False):  # Check if running as an executable
        base_path = sys._MEIPASS  # Temporary path where PyInstaller unpacks files
    else:
        base_path = os.path.dirname(__file__)  # Normal script execution path
    return os.path.join(base_path, filename)

# Load users.json
users_file = get_data_file_path("users.json")
with open(users_file, "r") as f:
    users_data = json.load(f)

# Load dsaprogress.json
dsa_file = get_data_file_path("dsaprogress.json")
with open(dsa_file, "r") as f:
    dsa_data = json.load(f)


# File paths
USER_FILE = "users.json"
PROGRESS_FILE = "dsaprogress.json"
EXCEL_FILE = "tasks.xlsx"

# Ensure JSON files exist
if not os.path.exists(USER_FILE):
    with open(USER_FILE, "w") as f:
        json.dump({"student": {"password": "student123"}, "admin": {"password": "admin123"}}, f)

if not os.path.exists(PROGRESS_FILE):
    with open(PROGRESS_FILE, "w") as f:
        json.dump({}, f)

# Global Style Configurations
FONT_TITLE = ("Arial", 18, "bold")
FONT_BODY = ("Arial", 10)
BTN_STYLE = {"font": ("Arial", 12), "borderwidth": 0, "relief": "flat", "cursor": "hand2"}

# Color Themes
THEMES = {
    "login": {"primary": "#4682b4", "secondary": "#5a9bd4", "bg": "#f0f8ff", "accent": "#b0c4de"},
    "dashboard": {"primary": "#6a5acd", "secondary": "#9370db", "bg": "#e6e6fa", "accent": "#d8bfd8"},
    "dsa": {"primary": "#2e8b57", "secondary": "#3cb371", "bg": "#f0fff0", "accent": "#98fb98"},
    "cpp": {"primary": "#b22222", "secondary": "#cd5c5c", "bg": "#fff0f5", "accent": "#ffb6c1"},
    "web": {"primary": "#20b2aa", "secondary": "#48d1cc", "bg": "#f0ffff", "accent": "#afeeee"},
    "admin": {"primary": "#34495e", "secondary": "#2c3e50", "bg": "#ecf0f1", "accent": "#bdc3c7"}
}

def create_decorations(window, theme):
    canvas = tk.Canvas(window, bg=theme["bg"], highlightthickness=0)
    canvas.place(relwidth=1, relheight=1)
    
    # Create geometric pattern
    for i in range(0, 150, 20):
        canvas.create_oval(i, 450, i+30, 480, fill=theme["accent"], outline="")
    for i in range(100, 500, 80):
        canvas.create_rectangle(i, 50, i+40, 70, fill=theme["accent"], outline="")
    canvas.create_polygon(400, 0, 500, 0, 500, 100, fill=theme["accent"], outline="")
    return canvas

def load_users():
    with open(USER_FILE, "r") as f:
        return json.load(f)

def save_users(users):
    with open(USER_FILE, "w") as f:
        json.dump(users, f, indent=4)

def save_progress(student_id, progress):
    all_progress = load_progress()
    all_progress[student_id] = progress
    with open(PROGRESS_FILE, "w") as f:
        json.dump(all_progress, f, indent=4)

def load_progress():
    with open(PROGRESS_FILE, "r") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {}

# Admin Panel
def open_admin_panel():
    theme = THEMES["admin"]
    admin_window = tk.Tk()
    admin_window.title("Admin Panel")
    admin_window.geometry("500x400")
    admin_window.configure(bg=theme["bg"])
    admin_window.resizable(False, False)

    create_decorations(admin_window, theme)

    header = tk.Frame(admin_window, bg=theme["primary"], height=80)
    header.pack(fill="x")
    tk.Label(header, text="Admin Panel", font=FONT_TITLE, 
            bg=theme["primary"], fg="white").pack(pady=15)

    main_frame = tk.Frame(admin_window, bg=theme["bg"])
    main_frame.pack(pady=20, padx=40, fill="both", expand=True)

    tk.Label(main_frame, text="Create New Student", font=FONT_BODY, 
            bg=theme["bg"], fg=theme["primary"]).pack(anchor="w", pady=(0, 5))

    # Username Entry
    tk.Label(main_frame, text="Username", font=FONT_BODY, 
            bg=theme["bg"], fg=theme["primary"]).pack(anchor="w", pady=(0, 5))
    new_username_entry = tk.Entry(main_frame, font=("Arial", 12), relief="flat", highlightthickness=1)
    new_username_entry.configure(highlightbackground=theme["accent"], highlightcolor=theme["primary"])
    new_username_entry.pack(fill="x", pady=(0, 15))

    # Password Entry
    tk.Label(main_frame, text="Password", font=FONT_BODY, 
            bg=theme["bg"], fg=theme["primary"]).pack(anchor="w", pady=(0, 5))
    new_password_entry = tk.Entry(main_frame, font=("Arial", 12), show="*", 
                                relief="flat", highlightthickness=1)
    new_password_entry.configure(highlightbackground=theme["accent"], highlightcolor=theme["primary"])
    new_password_entry.pack(fill="x", pady=(0, 25))

    def create_student():
        username = new_username_entry.get()
        password = new_password_entry.get()

        if not username or not password:
            messagebox.showerror("Error", "Username and password cannot be empty!")
            return

        users = load_users()
        if username in users:
            messagebox.showerror("Error", "Username already exists!")
            return

        users[username] = {"password": password}
        save_users(users)
        messagebox.showinfo("Success", "Student created successfully!")
        new_username_entry.delete(0, 'end')
        new_password_entry.delete(0, 'end')

    create_btn = tk.Button(main_frame, text="Create Student", command=create_student,
                          bg=theme["primary"], fg="white", **BTN_STYLE)
    create_btn.pack(fill="x", ipady=8)

    admin_window.mainloop()

# Login Window
def login_window():
    theme = THEMES["login"]
    login_win = tk.Tk()
    login_win.title("Login")
    login_win.geometry("500x500")
    login_win.configure(bg=theme["bg"])
    login_win.resizable(False, False)

    create_decorations(login_win, theme)

    header_frame = tk.Frame(login_win, bg=theme["primary"], height=120)
    header_frame.pack(fill="x")
    tk.Label(header_frame, text="Welcome Back", font=FONT_TITLE, 
            bg=theme["primary"], fg="white").pack(pady=20)

    main_frame = tk.Frame(login_win, bg=theme["bg"])
    main_frame.pack(pady=40, padx=40, fill="both", expand=True)

    tk.Label(main_frame, text="Username", font=FONT_BODY, 
            bg=theme["bg"], fg=theme["primary"]).pack(anchor="w", pady=(0, 5))
    username_entry = tk.Entry(main_frame, font=("Arial", 12), relief="flat", highlightthickness=1)
    username_entry.configure(highlightbackground=theme["accent"], highlightcolor=theme["primary"])
    username_entry.pack(fill="x", pady=(0, 15))

    tk.Label(main_frame, text="Password", font=FONT_BODY, 
            bg=theme["bg"], fg=theme["primary"]).pack(anchor="w", pady=(0, 5))
    password_entry = tk.Entry(main_frame, font=("Arial", 12), show="*", 
                            relief="flat", highlightthickness=1)
    password_entry.configure(highlightbackground=theme["accent"], highlightcolor=theme["primary"])
    password_entry.pack(fill="x", pady=(0, 25))

    def on_enter(e): login_btn.config(bg=theme["secondary"])
    def on_leave(e): login_btn.config(bg=theme["primary"])

    login_btn = tk.Button(main_frame, text="Sign In", font=("Arial", 12, "bold"), 
                         bg=theme["primary"], fg="white", cursor="hand2", relief="flat",
                         activebackground=theme["secondary"], activeforeground="white")
    login_btn.pack(fill="x", ipady=8)
    login_btn.bind("<Enter>", on_enter)
    login_btn.bind("<Leave>", on_leave)

    def login(event=None):
        users = load_users()
        username = username_entry.get()
        password = password_entry.get()

        # Check for admin login
        if username == "admin" and password == "admin123":
            login_win.destroy()
            open_admin_panel()
            return  # Exit the function after handling admin login

        # Check for normal user login
        if username in users and users[username]["password"] == password:
            login_win.destroy()
            open_student_dashboard(username)
        else:
            messagebox.showerror("Error", "Invalid Credentials")
            password_entry.delete(0, 'end')

    login_win.bind('<Return>', login)
    login_btn.config(command=login)
    login_win.mainloop()

# Student Dashboard
def open_student_dashboard(student_id):
    theme = THEMES["dashboard"]
    student_window = tk.Tk()
    student_window.title("Student Dashboard")
    student_window.geometry("600x600")
    student_window.configure(bg=theme["bg"])
    student_window.resizable(False, False)
    
    create_decorations(student_window, theme)
    
    header = tk.Frame(student_window, bg=theme["primary"], height=120)
    header.pack(fill="x")
    tk.Label(header, text=f"Welcome, {student_id}", font=FONT_TITLE, 
            bg=theme["primary"], fg="white").pack(pady=20)
    
    main_frame = tk.Frame(student_window, bg=theme["bg"])
    main_frame.pack(pady=20, padx=40, fill="both", expand=True)
    
    buttons = [
        ("C++ Course", lambda: open_cpp_section(student_window), THEMES["cpp"]["primary"]),
        ("DSA Tracker", lambda: open_dsa_tracker(student_window, student_id), THEMES["dsa"]["primary"]),
        ("Web Dev Course", lambda: open_web_dev_section(student_window), THEMES["web"]["primary"]),
        ("Database", lambda: open_youtube("https://youtu.be/dl00fOOYLOM?si=XA1qb7ntdnK9dJe3"), "#9b59b6"),
        ("OS", lambda: open_youtube("https://youtu.be/3obEP8eLsCw?si=1-Gv0kFrTVC0xkbB"), "#e67e22"),
        ("C Language", lambda: open_youtube("https://youtu.be/aZb0iu4uGwA?si=CnKPnVVUlhahUo_H"), "#3498db"),
        ("Python", lambda: open_youtube("https://youtu.be/UrsmFxEIp5k?si=HSDDGJb_MY-xTI05"), "#f1c40f")
    ]
    
    for text, cmd, color in buttons:
        btn = tk.Button(main_frame, text=text, command=cmd, 
                       bg=color, fg="white", **BTN_STYLE)
        btn.pack(fill="x", pady=5, ipady=8)
        btn.bind("<Enter>", lambda e, b=btn: b.config(bg=theme["secondary"]))
        btn.bind("<Leave>", lambda e, b=btn, c=color: b.config(bg=c))
    
    tk.Button(student_window, text="Logout", command=lambda: logout(student_window), 
             bg=theme["primary"], fg="white", **BTN_STYLE).pack(side="bottom", pady=10)
    
    student_window.mainloop()

# DSA Tracker with Mouse Wheel Scroll
def open_dsa_tracker(prev_window, student_id):
    theme = THEMES["dsa"]
    prev_window.destroy()
    
    tracker_window = tk.Tk()
    tracker_window.title("DSA Tracker")
    tracker_window.geometry("900x700")
    tracker_window.configure(bg=theme["bg"])
    tracker_window.resizable(False, False)
    
    create_decorations(tracker_window, theme)
    
    header = tk.Frame(tracker_window, bg=theme["primary"], height=80)
    header.pack(fill="x")
    tk.Label(header, text="DSA Progress Tracker", font=FONT_TITLE, 
            bg=theme["primary"], fg="white").pack(pady=15)
    
    tk.Button(tracker_window, text="← Back", command=lambda: go_back(tracker_window, student_id),
             bg=theme["primary"], fg="white", **BTN_STYLE).place(x=10, y=10)

    progress_frame = tk.Frame(tracker_window, bg=theme["bg"])
    progress_frame.pack(pady=20, padx=40, fill="x")

    # Load Excel with hyperlinks using openpyxl
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2):  # Skip header row
        topic = row[0].value
        question_name = row[1].value  # Extract the question name from the second column
        question_url = row[1].hyperlink.target if row[1].hyperlink else ""
        data.append([topic, question_name, question_url])
    
    df = pd.DataFrame(data, columns=["Topic", "Question", "URL"])
    progress = load_progress().get(student_id, {})

    progress_label = tk.Label(progress_frame, text="0/0 Questions Completed (0%)", 
                            bg=theme["bg"], fg=theme["primary"], font=("Arial", 12))
    progress_label.pack(pady=5)

    progress_bar_canvas = tk.Canvas(progress_frame, width=300, height=20, bg="white", highlightthickness=0)
    progress_bar_canvas.pack(pady=5)
    progress_line = progress_bar_canvas.create_rectangle(0, 0, 0, 20, fill=theme["primary"], outline="")

    frame = tk.Frame(tracker_window, bg=theme["bg"])
    frame.pack(fill="both", expand=True)

    canvas = tk.Canvas(frame, bg=theme["bg"], highlightthickness=0)
    scrollable_frame = tk.Frame(canvas, bg=theme["bg"])

    # Mouse wheel scrolling
    def _on_mousewheel(event):
        if platform.system() == "Windows":
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        elif platform.system() == "Darwin":
            canvas.yview_scroll(int(-1*event.delta), "units")
        else:
            if event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")

    def _bound_to_mousewheel(event):
        if platform.system() in ["Windows", "Darwin"]:
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        else:
            canvas.bind_all("<Button-4>", _on_mousewheel)
            canvas.bind_all("<Button-5>", _on_mousewheel)

    def _unbound_to_mousewheel(event):
        if platform.system() in ["Windows", "Darwin"]:
            canvas.unbind_all("<MouseWheel>")
        else:
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

    # Configure scrolling region
    def _configure_scroll_region(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
    
    scrollable_frame.bind("<Configure>", _configure_scroll_region)
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.bind("<Enter>", _bound_to_mousewheel)
    canvas.bind("<Leave>", _unbound_to_mousewheel)
    canvas.pack(side="left", fill="both", expand=True)

    def update_progress_ui():
        total = len(df)
        completed = sum(progress.values())
        percentage = (completed / total) * 100 if total > 0 else 0
        progress_label.config(text=f"{completed}/{total} Questions Completed ({percentage:.1f}%)")
        progress_width = (percentage / 100) * 300
        progress_bar_canvas.coords(progress_line, 0, 0, progress_width, 20)

    def toggle_done(question_url, label):
        progress[question_url] = not progress.get(question_url, False)
        save_progress(student_id, progress)
        label.config(text="✅" if progress[question_url] else "❌")
        update_progress_ui()

    for index, row in df.iterrows():
        topic = row["Topic"]
        question_name = row["Question"]
        question_url = row["URL"]
        done = progress.get(question_url, False)

        frame_question = tk.Frame(scrollable_frame, bg=theme["bg"])
        frame_question.pack(fill="x", padx=10, pady=2)

        # Display topic and question name
        label_text = f"{topic} - {question_name}"
        label = tk.Label(frame_question, text=label_text, 
                        fg=theme["primary"], bg=theme["bg"], cursor="hand2")
        label.pack(side="left", padx=5)
        if question_url:
            label.bind("<Button-1>", lambda e, url=question_url: webbrowser.open(url))

        done_label = tk.Label(frame_question, text="✅" if done else "❌", 
                            fg=theme["primary"], bg=theme["bg"], font=("Arial", 12), cursor="hand2")
        done_label.pack(side="right", padx=10)
        done_label.bind("<Button-1>", lambda e, q=question_url, l=done_label: toggle_done(q, l))

    update_progress_ui()
    tracker_window.mainloop()
# C++ Section
def open_cpp_section(prev_window):
    theme = THEMES["cpp"]
    prev_window.destroy()
    
    cpp_window = tk.Tk()
    cpp_window.title("C++ Course")
    cpp_window.geometry("500x400")
    cpp_window.configure(bg=theme["bg"])
    cpp_window.resizable(False, False)
    
    create_decorations(cpp_window, theme)
    
    header = tk.Frame(cpp_window, bg=theme["primary"], height=80)
    header.pack(fill="x")
    tk.Label(header, text="C++ Programming Course", font=FONT_TITLE, 
            bg=theme["primary"], fg="white").pack(pady=15)
    
    tk.Button(cpp_window, text="← Back", command=lambda: go_back(cpp_window),
             bg=theme["primary"], fg="white", **BTN_STYLE).place(x=10, y=10)

    main_frame = tk.Frame(cpp_window, bg=theme["bg"])
    main_frame.pack(pady=20, padx=40, fill="both", expand=True)

    links = [
        ("Learn C++ Language", "https://youtu.be/yGB9jhsEsr8"),
        ("STL Tutorial", "https://youtu.be/RRVYpIET_RU"),
        ("OOPs Concepts", "https://www.youtube.com/playlist?list=PLISTUNloqsz0z9JJJke7g7PxRLvy6How9")
    ]
    
    for text, url in links:
        btn = tk.Button(main_frame, text=text, command=lambda u=url: open_youtube(u),
                       bg=theme["primary"], fg="white", **BTN_STYLE)
        btn.pack(fill="x", pady=5, ipady=8)
        btn.bind("<Enter>", lambda e, b=btn: b.config(bg=theme["secondary"]))
        btn.bind("<Leave>", lambda e, b=btn: b.config(bg=theme["primary"]))
    
    cpp_window.mainloop()

# Web Development Section
def open_web_dev_section(prev_window):
    theme = THEMES["web"]
    prev_window.destroy()
    
    web_window = tk.Tk()
    web_window.title("Web Development")
    web_window.geometry("500x400")
    web_window.configure(bg=theme["bg"])
    web_window.resizable(False, False)
    
    create_decorations(web_window, theme)
    
    header = tk.Frame(web_window, bg=theme["primary"], height=80)
    header.pack(fill="x")
    tk.Label(header, text="Web Development Courses", font=FONT_TITLE, 
            bg=theme["primary"], fg="white").pack(pady=15)
    
    tk.Button(web_window, text="← Back", command=lambda: go_back(web_window),
             bg=theme["primary"], fg="white", **BTN_STYLE).place(x=10, y=10)

    main_frame = tk.Frame(web_window, bg=theme["bg"])
    main_frame.pack(pady=20, padx=40, fill="both", expand=True)

    links = [
        ("Full Course + Project (Hindi)", "https://www.youtube.com/playlist?list=PLu0W_9lII9agq5TrH9XLIKQvv0iaF2X3w"),
        ("HTML + CSS (English)", "https://youtu.be/G3e-cpL7ofc")
    ]
    
    for text, url in links:
        btn = tk.Button(main_frame, text=text, command=lambda u=url: open_youtube(u),
                       bg=theme["primary"], fg="white", **BTN_STYLE)
        btn.pack(fill="x", pady=5, ipady=8)
        btn.bind("<Enter>", lambda e, b=btn: b.config(bg=theme["secondary"]))
        btn.bind("<Leave>", lambda e, b=btn: b.config(bg=theme["primary"]))
    
    web_window.mainloop()

def open_youtube(url):
    webbrowser.open(url)

def logout(window):
    window.destroy()
    login_window()

def go_back(current_window, student_id="student"):
    current_window.destroy()
    open_student_dashboard(student_id)

if __name__ == "__main__":
    login_window()
